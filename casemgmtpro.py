import streamlit as st
import pandas as pd
import datetime
import pytz
from dateutil import parser
import requests, json, time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import os
from fpdf import FPDF

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

APP_NAME = "Case Pilot"
APP_SUB = "A Case Management Tool Developed and Created by JAY KISHAN SHARMA"
API_URL = "https://eciapi.akshit.me"
REQUIRED_COLUMNS = [
    "cino","type_name","case_no","reg_no","reg_year",
    "petparty_name","resparty_name","date_last_list","date_next_list",
    "purpose_name","disp_name","establishment_name","court_no_desg_name"
]
DEFAULT_CAUSELIST_COLUMNS = [
    "Previous Date",
    "court_no_desg_name",
    "Type",
    "Case Number/Year",
    "Parties",
    "Stage Today",
    "Next Date"
]
defaults = {
    "cases": pd.DataFrame(columns=REQUIRED_COLUMNS),
    "case_notes": {},
    "case_dossiers": {},
    "case_papers": {},
    "pinned_cases": set(),
    "reminders": [],
    "last_sync_date": None,
    "theme": "Dark",
    "auto_sync_time": datetime.time(17, 0),
    "api_key": "ECIAPI-xxxxxxxxxxxxxxxxxxxxxxxx",
    "billing_entries": [],
    "service_types": [
        "Hearing/Appearing Charges",
        "Witness Preparation Charges",
        "Drafting Legal Fees",
        "Notice Charges",
        "Legal Opinion Charges",
        "Registration Charges",
        "Filing Legal Fees",
        "Other"
    ],
    "causelist_columns": DEFAULT_CAUSELIST_COLUMNS.copy(),
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

def apply_theme():
    st.markdown(f'''
    <style>
        body, input, select, textarea {{
            font-family: 'Segoe UI', sans-serif !important;
            background-color: {'#121212' if st.session_state['theme']=="Dark" else '#FFFFFF'};
            color: {'#e0e0e0' if st.session_state['theme']=="Dark" else '#000000'};
        }}
        h1, h4 {{ text-align: center !important; }}
        .stTabs [role=tablist] {{ justify-content: center !important; }}
        table.dataframe th {{ text-align: center !important; }}
        button[title="Move Up"] {{margin-right: 5px;}}
        button[title="Move Down"] {{margin-left: 5px;}}
    </style>
    ''', unsafe_allow_html=True)

def today():
    return datetime.date.today()

def get_cases_on(d):
    return st.session_state.cases[st.session_state.cases["date_next_list"] == d]

def filter_next_30(keyword=None):
    start, end = today(), today() + datetime.timedelta(days=30)
    df = st.session_state.cases[
        (st.session_state.cases["date_next_list"] >= start) &
        (st.session_state.cases["date_next_list"] <= end)
    ]
    if keyword:
        df = df[df["purpose_name"].str.lower().str.contains(keyword, na=False)]
    out = prepare_display_df(df)
    cols = ["Next Date", "Case Number/Year", "Parties", "Stage Today", "Category"]
    existing = [c for c in cols if c in out.columns]
    return out[existing]

def assign_category(row):
    court_str = f"{row.get('establishment_name','')} {row.get('court_no_desg_name','')}".lower()
    case_str = f"{row.get('case_no','')} {row.get('reg_no','')} {row.get('type_name','')}".lower()

    case_type = str(row.get('type_name', '')).lower()
    case_no_full = f"{row.get('case_no','')} {row.get('reg_no','')}".lower()

    reg_no = row.get("reg_no")
    try:
        case_number = int(reg_no) if reg_no is not None else None
    except ValueError:
        case_number = None

    # ----- Threshold rules -----
    if case_number is not None:
        # CC / C.C / Crime cases
        if any(x in case_type for x in ["cc", "c.c", "crime", "criminal case"]):
            if case_number > 50000:
                return "FC/MAYO/COM/CONS/DRT/OUT"
            else:
                return "ACMM/ACJM/MMTC"

        # CRL.A / CRL.RP cases
        if any(x in case_type for x in ["crl.a", "crl.rp", "crl.r.p"]):
            if case_number > 20000:
                return "FC/MAYO/COM/CONS/DRT/OUT"
            else:
                return "CCC/S/SCCH/MACT"

        # S.C / SC cases
        if case_type in ["s.c", "sc"]:
            if case_number > 15000:
                return "FC/MAYO/COM/CONS/DRT/OUT"
            else:
                return "CCC/S/SCCH/MACT"


    if "mayo" in court_str or "mayohall" in court_str:
        return "FC/MAYO/COM/CONS/DRT/OUT"

    if ("commercial" in case_str or
        any(x in case_str for x in ["com os","com.os","com ex","com.ex"]) or
        case_str.startswith("com")):
        return "FC/MAYO/COM/CONS/DRT/OUT"

    if "magistrate" in court_str:
        if "mayo" in court_str or "mayohall" in court_str:
            return "FC/MAYO/COM/CONS/DRT/OUT"
        return "ACMM/ACJM/MMTC"

    if any(x in court_str for x in ["city civil","sessions","small causes","scch","mact","rural"]):
        return "CCC/S/SCCH/MACT"

    return "FC/MAYO/COM/CONS/DRT/OUT"

def load_cases(file):
    try:
        raw = file.read().decode("utf-8")
        data = json.loads(raw)
        if isinstance(data, dict) and all(isinstance(v, dict) for v in data.values()):
            data = list(data.values())
        if data and all(isinstance(x, str) for x in data):
            data = [json.loads(x) for x in data]
        df = pd.DataFrame(data)
        for c in REQUIRED_COLUMNS:
            if c not in df.columns:
                df[c] = None
        for c in ["date_last_list", "date_next_list"]:
            df[c] = pd.to_datetime(df[c], dayfirst=True, errors="coerce").dt.date
        st.session_state.cases = df[REQUIRED_COLUMNS]
        st.success(f"Loaded {len(df)} cases.")
    except Exception as e:
        st.error(f"Failed loading cases: {e}")

def prepare_display_df(df):
    out = df.copy()
    out["Previous Date"] = out["date_last_list"].apply(lambda x: x.strftime("%d.%m.%Y") if pd.notna(x) else "")
    out["Next Date"] = out["date_next_list"].apply(lambda x: x.strftime("%d.%m.%Y") if pd.notna(x) else "")
    out["Case Number/Year"] = out["reg_no"].astype(str) + "/" + out["reg_year"].astype(str)
    out["Parties"] = out["petparty_name"].fillna("") + " v. " + out["resparty_name"].fillna("")
    out["Stage Today"] = out["purpose_name"] if "purpose_name" in out.columns else ""
    out["Type"] = out["type_name"].fillna("")
    out["Category"] = out.apply(assign_category, axis=1)
    return out

def export_cause_list_excel_categorized(df, selected_columns, filename="Cause_List"):
    wb = Workbook()
    ws = wb.active
    ws.title = "CauseList"
    max_col = len(selected_columns)

    font_header = Font(size=20, bold=True)
    font_footer = Font(size=20, italic=True)
    font_data = Font(size=20)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1)
    cell.value = APP_NAME
    cell.font = font_header
    cell.alignment = align_center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    cell = ws.cell(row=2, column=1)
    cell.value = APP_SUB
    cell.font = font_header
    cell.alignment = align_center

    ws.append([])  # blank line

    # Add column headers
    ws.append(selected_columns)
    for cell in ws[ws.max_row]:
        cell.font = font_header
        cell.alignment = align_center

    # Add data rows with empty Next Date
    for _, row in df.iterrows():
        row_data = []
        for col in selected_columns:
            if col == "Next Date":
                row_data.append("")  # Always empty
            else:
                val = row.get(col, "")
                if isinstance(val, str) and len(val) > 50:
                    val = val[:47] + "..."
                row_data.append(val)
        ws.append(row_data)

    # Style data rows
    for row_cells in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row_cells:
            cell.font = font_data
            cell.alignment = align_left_wrap

    # Additional recommended notes (3 rows)
    ws.append([])
    ws.append(["Additional Category Notes:"])
    ws.append(["*Cases advanced/listed but not appearing in cause list; Certified copies, Compliance, Office cases, Client appearances, Follow-ups, etc."])
    for r in range(ws.max_row-2, ws.max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_footer
            cell.alignment = align_left_wrap

    # Footer below notes with spacing
    footer_row = ws.max_row + 3
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=max_col)
    

    # Page setup
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    ws.oddFooter.center.text = f"{APP_NAME} - {APP_SUB}"
    ws.oddFooter.center.size = 18
    ws.oddFooter.center.font = "Arial"
    ws.oddFooter.center.color = "000000"


    # Adjust column width
    for i, col in enumerate(selected_columns, 1):
        max_len = len(col) + 2
        for cell in ws[get_column_letter(i)]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 5, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    st.download_button(f"Download {filename} Excel", data=buffer,
                       file_name=f"{filename}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def generate_cause_list_pdf(df, selected_columns, filename="Cause_List.pdf", category_name=None):
    pdf = FPDF(unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, APP_NAME, ln=True, align="C")
    pdf.set_font("Arial", '', 12)
    pdf.cell(0, 8, APP_SUB, ln=True, align="C")
    pdf.ln(6)

    # ---- CATEGORY HEADER ----
    if category_name:
        pdf.set_font("Arial", 'B', 13)
        pdf.cell(0, 8, category_name, ln=True, align="C")
        pdf.ln(4)  # some spacing before table

    # --- Column widths ---
    base_widths = [25, 35, 20, 35, 55, 35, 25]
    if len(selected_columns) != len(base_widths):
        base_widths = [1] * len(selected_columns)

    available = pdf.w - pdf.l_margin - pdf.r_margin
    scale = available / sum(base_widths)
    col_widths = [w * scale for w in base_widths]

    # Header drawer
    def draw_header():
        pdf.set_font("Arial", 'B', 10)
        header_h = 8
        x = pdf.l_margin
        y = pdf.get_y()
        for i, col in enumerate(selected_columns):
            pdf.rect(x, y, col_widths[i], header_h)
            pdf.set_xy(x, y)
            pdf.multi_cell(col_widths[i], header_h, col, border=0, align="C")
            x += col_widths[i]
        pdf.ln(header_h)

    draw_header()

    # --- Rows (unchanged) ---
    line_h = 5
    pdf.set_font("Arial", '', 9)

    for _, row in df.iterrows():
        cell_lines = []
        max_lines = 1
        for i, col in enumerate(selected_columns):
            text = "" if col == "Next Date" else str(row.get(col, "") or "")
            lines = pdf.multi_cell(col_widths[i], line_h, text, border=0, align="L", split_only=True)
            if not lines:
                lines = [""]
            cell_lines.append(lines)
            if len(lines) > max_lines:
                max_lines = len(lines)

        row_h = max_lines * line_h

        if pdf.will_page_break(row_h):
            pdf.add_page()
            if category_name:
                pdf.set_font("Arial", 'B', 13)
                pdf.cell(0, 8, category_name, ln=True, align="C")
                pdf.ln(4)
            draw_header()

        y = pdf.get_y()
        x = pdf.l_margin
        for i, lines in enumerate(cell_lines):
            pdf.rect(x, y, col_widths[i], row_h)
            pdf.set_xy(x, y)
            pdf.multi_cell(col_widths[i], line_h, "\n".join(lines), border=0, align="L")
            x += col_widths[i]
        pdf.set_xy(pdf.l_margin, y + row_h)

    # Footer remains same...
    pdf.ln(5)
    note = "*Cases advanced/listed but not appearing in cause list; Certified copies, Compliance, Office cases, Client appearances, Follow-ups, etc."
    if pdf.will_page_break(18):
        pdf.add_page()
        if category_name:
            pdf.set_font("Arial", 'B', 13)
            pdf.cell(0, 8, category_name, ln=True, align="C")
            pdf.ln(4)
        draw_header()
    pdf.set_font("Arial", 'I', 9)
    pdf.multi_cell(0, 6, note)

    pdf.ln(6)
    pdf.set_font("Arial", 'I', 10)
    pdf.cell(0, 8, APP_NAME, align="C", ln=True)
    pdf.cell(0, 8, APP_SUB, align="C")

    pdf_output = bytes(pdf.output(dest="S"))
    st.download_button("Download Cause List PDF", pdf_output, file_name=filename, mime="application/pdf")




def cause_list_tab():
    st.subheader("Cause Lists - Select Date and View by Category")

    if st.session_state.cases.empty:
        st.info("No cases loaded.")
        return

    date_choice = st.radio("View Cause List For:", ["Today", "Tomorrow"])
    selected_date = today() if date_choice == "Today" else today() + datetime.timedelta(days=1)
    df = st.session_state.cases[st.session_state.cases["date_next_list"] == selected_date]

    if df.empty:
        st.info(f"No cases listed for {date_choice}.")
        return

    df_prepared = prepare_display_df(df)

    df_prepared["Category"] = df_prepared.apply(assign_category, axis=1)

    df_prepared = df_prepared.rename(columns={"court_no_desg_name": "Court Hall"})

    display_columns = ["Previous Date", "Court Hall", "Type",
                       "Case Number/Year", "Parties", "Stage Today", "Next Date"]

    categories = ["CCC/S/SCCH/MACT", "ACMM/ACJM/MMTC", "FC/MAYO/COM/CONS/DRT/OUT"]

    for cat in categories:
        cat_cases = df_prepared[df_prepared["Category"] == cat]
        if cat_cases.empty:
            continue
        st.markdown(f"### {cat}")
        st.dataframe(cat_cases[display_columns], use_container_width=True)

        if st.button(f"Export {cat} Cause List to Excel"):
            export_cause_list_excel_categorized(cat_cases, display_columns, f"{date_choice}_Cause_List_{cat.replace('/', '_')}")
        if st.button(f"Export {cat} Cause List to PDF"):
            generate_cause_list_pdf(cat_cases, display_columns, f"{date_choice}_Cause_List_{cat.replace('/', '_')}.pdf", category_name=cat)


def case_papers_tab():
    st.subheader("Case Papers Organisation")
    upload_folder = "uploaded_case_docs"
    os.makedirs(upload_folder, exist_ok=True)
    if st.session_state.cases.empty:
        st.info("No cases loaded.")
        return

    display_case_list = []
    for _, row in st.session_state.cases.iterrows():
        case_type = row.get('type_name', 'N/A')
        reg_no_year = f"{row['reg_no']}/{row['reg_year']}" if pd.notna(row['reg_no']) and pd.notna(row['reg_year']) else "N/A"
        parties = f"{row['petparty_name']} v. {row['resparty_name']}".strip()
        display_case_list.append(f"{row['cino']} - {case_type} - {reg_no_year} - {parties}")

    sel_display = st.selectbox("Assign Documents to Case", options=display_case_list)
    sel_cino = sel_display.split(" - ")[0]

    doc_type = st.selectbox("Document Type", ["Pleading", "Evidence", "Order Copy", "Other"])
    custom_doc_name = st.text_input("Custom Document Name (optional)")

    uploaded_files = st.file_uploader("Upload Documents", type=["pdf", "docx", "jpg", "png"], accept_multiple_files=True)

    if uploaded_files and st.button("Save Uploaded Documents"):
        for f in uploaded_files:
            base_filename = custom_doc_name.strip() or f.name
            safe_filename = f"{sel_cino}_{doc_type}_{base_filename}"
            file_path = os.path.join(upload_folder, safe_filename)
            with open(file_path, "wb") as out_file:
                out_file.write(f.read())
            if sel_cino not in st.session_state.case_papers:
                st.session_state.case_papers[sel_cino] = []
            st.session_state.case_papers[sel_cino].append({
                "doc_type": doc_type,
                "custom_doc_name": base_filename,
                "original_file_name": f.name,
                "path": file_path,
            })
        st.success(f"Uploaded {len(uploaded_files)} documents for case {sel_cino}.")

    search_term = st.text_input("Search Uploaded Documents by case number, parties, type or document name", value="")
    st.markdown("### Uploaded Documents")
    documents_to_show = []
    for cino, docs in st.session_state.case_papers.items():
        case_row = st.session_state.cases[st.session_state.cases["cino"] == cino]
        if case_row.empty:
            continue
        party_str = f"{case_row.iloc[0]['petparty_name']} v. {case_row.iloc[0]['resparty_name']}".lower()
        reg_no_year = f"{case_row.iloc[0]['reg_no']}/{case_row.iloc[0]['reg_year']}"
        case_type = case_row.iloc[0].get('type_name', '').lower()
        searchable_str = f"{cino} {party_str} {reg_no_year} {case_type}".lower()
        for doc in docs:
            combined_search = f"{searchable_str} {doc['custom_doc_name'].lower()} {doc['original_file_name'].lower()} {doc['doc_type'].lower()}"
            if search_term.strip().lower() in combined_search:
                documents_to_show.append((cino, party_str, reg_no_year, case_type, doc))

    if documents_to_show:
        for cino, parties, reg_no_year, case_type, doc in documents_to_show:
            st.write(f"**Case:** {cino} - {case_type} - {reg_no_year} - {parties}")
            st.write(f"- Document Type: {doc['doc_type']}")
            st.write(f"- Document Name: {doc['custom_doc_name']} (Original: {doc['original_file_name']})")
            st.write(f"- Saved Path: {doc['path']}")
            st.markdown("---")
    else:
        st.write("No documents match the search criteria.")

def billing_tab():
    st.subheader("Billing & Time Tracking")
    with st.expander("Manage Service/Billing Categories"):
        st.write("Current Categories:")
        categories = st.session_state.service_types
        for idx, cat in enumerate(categories):
            cols = st.columns([7, 1])
            cols[0].write(cat)
            if cols[1].button("Remove", key=f"remove_service_{idx}"):
                categories.pop(idx)
                st.session_state.service_types = categories
                st.experimental_rerun()
        new_cat = st.text_input("Add New Billing Category")
        if st.button("Add Category") and new_cat.strip():
            if new_cat.strip() not in categories:
                categories.append(new_cat.strip())
                st.session_state.service_types = categories
                st.success(f"Added new billing category: {new_cat.strip()}")
                st.experimental_rerun()
    if "billing_entries" not in st.session_state:
        st.session_state.billing_entries = []
    case_options = ["General"] + list(st.session_state.cases["cino"]) if not st.session_state.cases.empty else ["General"]
    sel_case = st.selectbox("Select Case (or General)", case_options)
    with st.form("Add Billing Entry", clear_on_submit=True):
        entry_date = st.date_input("Date of Billing", datetime.date.today())
        service_type = st.selectbox("Billing Category", st.session_state.service_types)
        description = st.text_area("Description / Notes", height=70)
        fee_type = st.radio("Fee Type", ["Service Fee (Fixed)", "Time Based (Hourly)"])
        amount = st.number_input("Amount (INR)", min_value=0.0, format="%.2f")
        time_spent = 0.0
        if fee_type == "Time Based (Hourly)":
            time_spent = st.number_input("Time Spent (hours)", min_value=0.0, format="%.2f")
        submitted = st.form_submit_button("Add Billing Entry")
        if submitted:
            if amount <= 0:
                st.error("Amount must be greater than zero.")
            else:
                billing_record = {
                    "case": sel_case,
                    "date": entry_date.strftime("%d.%m.%Y"),
                    "service_type": service_type,
                    "description": description.strip(),
                    "fee_type": fee_type,
                    "amount": amount,
                    "time_spent": time_spent,
                }
                st.session_state.billing_entries.append(billing_record)
                st.success("Billing entry added.")
    df = pd.DataFrame(st.session_state.billing_entries)
    edited_df = st.data_editor(df, key="billing_editor", num_rows="dynamic")
    st.session_state.billing_entries = edited_df.to_dict("records")
    st.markdown("---")
    st.subheader("Billing Entries Summary")
    filter_case = st.selectbox("Filter by Case", ["All"] + case_options)
    filtered_entries = st.session_state.billing_entries
    if filter_case != "All":
        filtered_entries = [e for e in filtered_entries if e["case"] == filter_case]
    if filtered_entries:
        df_filtered = pd.DataFrame(filtered_entries)
        total_amount = df_filtered["amount"].sum()
        total_time = df_filtered["time_spent"].sum() if "time_spent" in df_filtered.columns else 0
        st.write(f"**Total Amount: ₹{total_amount:,.2f}**")
        if total_time > 0:
            st.write(f"**Total Time Spent: {total_time:.2f} hours**")
        display_df = df_filtered[["date","case","service_type","description","fee_type","amount","time_spent"]]
        st.dataframe(display_df, height=350)
        def export_billing_excel(dataframe):
            wb = Workbook()
            ws = wb.active
            ws.title = "Billing"
            ws.append(["Date","Case","Billing Category","Description", "Fee Type", "Amount (INR)", "Time Spent (hours)"])
            for r in dataframe.itertuples(index=False):
                ws.append(list(r))
            for i, col in enumerate(ws.columns, 1):
                max_len = 0
                for cell in col:
                    if hasattr(cell, "alignment"):
                        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = min(max_len + 5, 50)
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        if st.button("Export Billing Data to Excel"):
            buf = export_billing_excel(display_df)
            st.download_button("Download Billing Excel", buf, "billing.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("No billing entries to display.")

def judge_analytics_tab():
    st.subheader("Judge & Court Analytics")
    if st.session_state.cases.empty:
        st.info("No case data loaded.")
        return
    judge_counts = st.session_state.cases["court_no_desg_name"].value_counts()
    st.write("Cases per Judge/Court:")
    st.bar_chart(judge_counts)
    stage_counts = st.session_state.cases["purpose_name"].value_counts().head(10)
    fig, ax = plt.subplots()
    ax.bar(stage_counts.index, stage_counts.values)
    ax.set_xticklabels(stage_counts.index, rotation=45, ha="right")
    ax.set_title("Top 10 Hearing Stages")
    st.pyplot(fig)

def roll_cases_to_tomorrow():
    today_date = today()
    tomorrow = today_date + datetime.timedelta(days=1)
    df = st.session_state.cases.copy()
    updated_rows = 0
    for idx, row in df.iterrows():
        if row["date_next_list"] == today_date:
            updated_data = fetch_case_api(row["cino"])
            if updated_data and updated_data["date_next_list"] == tomorrow:
                df.at[idx, "date_last_list"] = row["date_next_list"]
                df.at[idx, "date_next_list"] = tomorrow
                updated_rows += 1
    st.session_state.cases = df
    if updated_rows > 0:
        st.info(f"Rolled {updated_rows} case(s) to Tomorrow's Cause List.")

def fetch_case_api(cino):
    try:
        r = requests.get(f"{API_URL}/case-status/{cino}", headers={"Authorization": f"Bearer {st.session_state['api_key']}"}, timeout=20)
        if r.status_code != 200:
            return None
        data = r.json()
        return {
            "date_last_list": parser.parse(data.get("date_last_list")).date() if data.get("date_last_list") else None,
            "date_next_list": parser.parse(data.get("date_next_list")).date() if data.get("date_next_list") else None,
            "purpose_name": data.get("purpose_name", "")
        }
    except:
        return None

def update_cases_api(only_today=False):
    df = st.session_state.cases.copy()
    target = df if not only_today else get_cases_on(today())
    if target.empty:
        st.info("No cases to update.")
        return
    bar = st.progress(0)
    updated = 0
    for i, idx in enumerate(target.index):
        cino = df.at[idx, "cino"]
        if not cino:
            continue
        upd = fetch_case_api(cino)
        if upd:
            df.at[idx, "date_last_list"] = df.at[idx, "date_next_list"]
            if upd["date_next_list"]:
                df.at[idx, "date_next_list"] = upd["date_next_list"]
            if upd["purpose_name"]:
                df.at[idx, "purpose_name"] = upd["purpose_name"]
            updated += 1
        bar.progress((i + 1) / len(target))
        time.sleep(0.2)
    st.session_state.cases = df
    st.session_state.last_sync_date = today()
    st.success(f"Updated {updated} cases.")

def main():
    st.set_page_config(page_title=APP_NAME, layout="wide")
    apply_theme()
    st.markdown(f"<div style='text-align:center'><h1>{APP_NAME}</h1><h4>{APP_SUB}</h4></div>", unsafe_allow_html=True)
    tabs = st.tabs([
        "🏠 Dashboard", "📋 Master List", "📅 Cause Lists", "📌 Pinned Cases", "📝 Case Details",
        "⏰ Reminders", "🔍 Search", "🔄 API Sync", "📅 Calendar", "📊 Analytics",
        "💾 Backup", "⚙️ Settings", "📂 Case Papers Org", "💼 Billing", "📈 Judge Analytics"
    ])
    (tab_dashboard, tab_master, tab_cause, tab_pinned, tab_details,
    tab_reminders, tab_search, tab_api, tab_calendar, tab_analytics,
    tab_backup, tab_settings, tab_casepapers, tab_billing, tab_judgeanalytics) = tabs

    now_ist = datetime.datetime.now(pytz.timezone("Asia/Kolkata")).time()
    if now_ist.hour == 17 and now_ist.minute < 5:
        roll_cases_to_tomorrow()

    with tab_dashboard:
        col1, col2 = st.columns([2, 1])
        with col1:
            st.metric("Total Cases", len(st.session_state.cases))
            st.metric("Today’s Cases", len(get_cases_on(today())))
            st.metric("Tomorrow’s Cases", len(get_cases_on(today() + datetime.timedelta(days=1))))
        with col2:
            critical_cases = get_cases_on(today())
            critical_cases = prepare_display_df(critical_cases)
            crit_filter = critical_cases["Stage Today"].str.lower().str.contains("argument|evidence|order|judgment|hearing", na=False)
            critical_cases = critical_cases[crit_filter]
            st.subheader("Critical Matters Today")
            st.dataframe(critical_cases, use_container_width=True)
        st.subheader("Next 30 Days Overview")
        st.write("**All Hearings:**")
        st.dataframe(filter_next_30(), use_container_width=True)
        st.write("**Written Statement:**")
        st.dataframe(filter_next_30("written statement"), use_container_width=True)
        st.write("**Evidence / Cross Examination:**")
        st.dataframe(filter_next_30("evidence|cross"), use_container_width=True)
        st.write("**Arguments:**")
        st.dataframe(filter_next_30("argument"), use_container_width=True)
        st.write("**Orders:**")
        st.dataframe(filter_next_30("order"), use_container_width=True)
        st.write("**Judgments:**")
        st.dataframe(filter_next_30("judgment"), use_container_width=True)

    with tab_master:
        f = st.file_uploader("Upload myCases.txt", type=["txt"])
        if f:
            load_cases(f)
        if not st.session_state.cases.empty:
            disp = prepare_display_df(st.session_state.cases)
            st.dataframe(disp, use_container_width=True)
            export_cause_list_excel_categorized(disp, st.session_state.causelist_columns, "Master_List")

    with tab_cause:
        cause_list_tab()

    with tab_pinned:
        st.subheader("Pinned Cases")
        if st.session_state.pinned_cases:
            pins = [st.session_state.cases[st.session_state.cases["cino"] == cino].iloc[0]
                    for cino in st.session_state.pinned_cases
                    if cino in st.session_state.cases["cino"].values]
            disp = prepare_display_df(pd.DataFrame(pins))
            st.dataframe(disp, use_container_width=True)
        else:
            st.info("No pinned cases.")

    with tab_details:
        if st.session_state.cases.empty:
            st.info("Load cases first.")
        else:
            sel = st.selectbox("Select Case (CINO)", st.session_state.cases["cino"])
            row_df = st.session_state.cases[st.session_state.cases["cino"] == sel]
            if not row_df.empty:
                row_dict = row_df.iloc[0].to_dict()
                detail_df = pd.DataFrame(list(row_dict.items()), columns=["Field", "Value"])
                st.table(detail_df)
            note_list = st.session_state.case_notes.get(sel, [])
            st.subheader("Personal Notes")
            # Show existing notes with timestamps
            if note_list:
                for idx, note_entry in enumerate(note_list):
                    st.markdown(f"**[{note_entry['date']}]**: {note_entry['text']}")
                    st.markdown("---")
            else:
                st.write("No personal notes for this case.")
            new_note_text = st.text_area("Add New Personal Note")
            new_note_date = st.date_input("Date for this Note", value=today())
            if st.button("Add Note"):
                if new_note_text.strip():
                    note_list.append({"date": new_note_date.strftime("%d.%m.%Y"), "text": new_note_text.strip()})
                    st.session_state.case_notes[sel] = note_list
                    st.success("Note added.")
                else:
                    st.error("Note text cannot be empty.")

            st.subheader("Case Dossier Timeline")
            timeline = st.session_state.case_dossiers.get(sel, [])
            if timeline:
                st.table(pd.DataFrame(timeline))
            else:
                st.write("No timeline available.")

    with tab_reminders:
        st.subheader("Task & Deadline Tracker")
        txt = st.text_input("Task Description")
        due = st.date_input("Due Date", today())
        if st.button("Add Task") and txt:
            st.session_state.reminders.append({"text": txt, "due": due})
            st.success("Task added.")
        if st.session_state.reminders:
            rem_df = pd.DataFrame(st.session_state.reminders).sort_values("due")
            rem_df['due'] = rem_df['due'].apply(lambda x: x.strftime("%d.%m.%Y") if isinstance(x, (datetime.date, datetime.datetime)) else str(x))
            st.table(rem_df)
        else:
            st.info("No tasks/reminders.")

    with tab_search:
        st.subheader("Search Cases")
        d = st.date_input("Filter by Hearing Date (optional)", value=None)
        term = st.text_input("Global Search Term")
        df = st.session_state.cases
        if d:
            df = df[df["date_next_list"] == d]
        if term:
            df = df[df.apply(lambda row: term.lower() in str(row).lower(), axis=1)]
        disp = prepare_display_df(df)
        st.dataframe(disp, use_container_width=True)

    with tab_api:
        st.subheader("API Sync")
        if st.button("Sync Today's Cases"):
            update_cases_api(only_today=True)
        if st.button("Sync All Cases"):
            update_cases_api(only_today=False)
        last_sync = st.session_state.get("last_sync_date", "Never")
        st.write(f"Last sync date: {last_sync}")

    with tab_calendar:
        st.subheader("Hearing Calendar")
        cal_df = st.session_state.cases.dropna(subset=["date_next_list"]).sort_values("date_next_list")
        st.dataframe(prepare_display_df(cal_df), use_container_width=True)
        if st.button("Export Calendar to Excel"):
            export_cause_list_excel_categorized(prepare_display_df(cal_df), st.session_state.causelist_columns, "Calendar_View")

    with tab_analytics:
        st.subheader("Analytics Overview")
        if st.session_state.cases.empty:
            st.info("No data loaded.")
        else:
            cats = st.session_state.cases.apply(assign_category, axis=1).value_counts()
            fig1, ax1 = plt.subplots()
            ax1.pie(cats, labels=cats.index, autopct='%1.1f%%')
            st.pyplot(fig1)
            st.subheader("Top 10 Hearing Stages")
            stage_counts = st.session_state.cases["purpose_name"].value_counts().head(10)
            fig2, ax2 = plt.subplots()
            ax2.bar(stage_counts.index, stage_counts.values)
            ax2.set_xticklabels(stage_counts.index, rotation=45, ha="right")
            st.pyplot(fig2)

    with tab_backup:
        st.subheader("Backup & Restore")
        if st.button("Backup Data"):
            backup = {
                "cases": st.session_state.cases.to_dict(),
                "case_notes": st.session_state.case_notes,
                "case_dossiers": st.session_state.case_dossiers,
                "case_papers": st.session_state.case_papers,
                "pinned_cases": list(st.session_state.pinned_cases),
                "reminders": st.session_state.reminders,
                "billing_entries": st.session_state.billing_entries,
                "service_types": st.session_state.service_types,
                "causelist_columns": st.session_state.causelist_columns,
                "last_sync_date": st.session_state.last_sync_date
            }
            buf = BytesIO(json.dumps(backup).encode("utf-8"))
            st.download_button("Download Backup JSON", buf, file_name="backup.json", mime="application/json")
        restore = st.file_uploader("Restore from Backup", type=["json"])
        if restore:
            data = json.load(restore)
            st.session_state.cases = pd.DataFrame(data.get("cases", {}))
            st.session_state.case_notes = data.get("case_notes", {})
            st.session_state.case_dossiers = data.get("case_dossiers", {})
            st.session_state.case_papers = data.get("case_papers", {})
            st.session_state.pinned_cases = set(data.get("pinned_cases", []))
            st.session_state.reminders = data.get("reminders", [])
            st.session_state.billing_entries = data.get("billing_entries", [])
            st.session_state.service_types = data.get("service_types", [])
            st.session_state.causelist_columns = data.get("causelist_columns", DEFAULT_CAUSELIST_COLUMNS.copy())
            st.session_state.last_sync_date = data.get("last_sync_date")
            st.success("Backup restored.")

    with tab_settings:
        st.subheader("Settings")
        st.session_state["theme"] = st.radio("Theme", ["Dark", "Light"], index=0 if st.session_state["theme"] == "Dark" else 1)
        st.session_state["auto_sync_time"] = st.time_input("Daily auto-sync time", st.session_state["auto_sync_time"])
        st.session_state["api_key"] = st.text_input("API Key", value=st.session_state["api_key"])

    with tab_casepapers:
        case_papers_tab()

    with tab_billing:
        billing_tab()

    with tab_judgeanalytics:
        judge_analytics_tab()

if __name__ == "__main__":
    main()
